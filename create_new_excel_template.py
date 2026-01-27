"""
創建新的 Excel 模板，包含所有需要的欄位
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_new_template(output_excel='learning_bubble_template_new.xlsx'):
    """創建包含所有欄位的新 Excel 模板"""
    
    print('📋 創建新的 Excel 模板...\n')
    
    # 定義所有工作表的欄位
    sheets_data = {
        'UI_SEGMENTS': {
            'columns': [
                'segmentId',      # ⭐ 必要
                'title',          # ⭐ 必要
                'order',          # ⭐ 必要
                'mode',           # ⭐ 必要
                'published',      # ⭐ 必要
                'tag',            # 選填
            ],
            'descriptions': {
                'segmentId': '區段 ID（唯一識別碼）',
                'title': '區段標題',
                'order': '排序順序（數字，越小越前面）',
                'mode': '模式（如：library, featured 等）',
                'published': '是否發布（true/false）',
                'tag': '標籤（選填）',
            }
        },
        'TOPICS': {
            'columns': [
                'topicId',        # ⭐ 必要
                'title',          # ⭐ 必要
                'published',      # ⭐ 必要
                'order',          # ⭐ 必要
                'tags',           # 選填（分號分隔）
                'bubbleImageUrl', # 選填
                'bubbleStorageFile', # 選填
                'bubbleGradStart', # 選填
                'bubbleGradEnd',   # 選填
            ],
            'descriptions': {
                'topicId': '主題 ID（唯一識別碼）',
                'title': '主題標題',
                'published': '是否發布（true/false）',
                'order': '排序順序（數字，越小越前面）',
                'tags': '標籤（多個用分號 ; 分隔）',
                'bubbleImageUrl': '泡泡圖片 URL',
                'bubbleStorageFile': '泡泡圖片儲存檔案路徑',
                'bubbleGradStart': '漸層起始顏色',
                'bubbleGradEnd': '漸層結束顏色',
            }
        },
        'PRODUCTS': {
            'columns': [
                'productId',      # ⭐ 必要
                'topicId',        # ⭐ 必要
                'level',          # ⭐ 必要
                'title',          # 選填（可自動生成：topicId + level）
                'titleLower',     # 選填（可自動生成：title 小寫）
                'order',          # 選填（預設 0）
                'type',           # 選填
                'published',      # 選填（預設 true）
                'levelGoal',      # 選填
                'levelBenefit',   # 選填
                'anchorGroup',    # 選填
                'version',        # 選填
                'coverImageUrl',  # 選填
                'coverStorageFile', # 選填
                'itemCount',      # 選填（數字）
                'wordCountAvg',   # 選填（數字）
                'pushStrategy',   # 選填
                'sourceType',     # 選填
                'source',         # 選填
                'sourceUrl',      # 選填
                'spec1Label',     # 選填
                'spec2Label',     # 選填
                'spec3Label',     # 選填
                'spec4Label',     # 選填
                'spec1Icon',      # 選填
                'spec2Icon',      # 選填
                'spec3Icon',      # 選填
                'spec4Icon',      # 選填
                'trialMode',      # 選填
                'trialLimit',     # 選填（數字，預設 3）
                'releaseAtMs',    # ⭐ 新增：發布時間戳（毫秒）
                'createdAtMs',   # ⭐ 新增：建立時間戳（毫秒）
            ],
            'descriptions': {
                'productId': '產品 ID（唯一識別碼）',
                'topicId': '所屬主題 ID',
                'level': '等級（如：L1, L2 等）',
                'title': '產品標題（留空則自動生成：topicId + level）',
                'titleLower': '產品標題小寫（留空則自動生成）',
                'order': '排序順序（數字，越小越前面，預設 0）',
                'type': '產品類型',
                'published': '是否發布（true/false，預設 true）',
                'levelGoal': '等級目標描述',
                'levelBenefit': '等級效益描述',
                'anchorGroup': '錨點群組',
                'version': '版本號',
                'coverImageUrl': '封面圖片 URL',
                'coverStorageFile': '封面圖片儲存檔案路徑',
                'itemCount': '內容項目數量',
                'wordCountAvg': '平均字數',
                'pushStrategy': '推播策略（如：seq）',
                'sourceType': '來源類型',
                'source': '來源',
                'sourceUrl': '來源 URL',
                'spec1Label': '規格 1 標籤',
                'spec2Label': '規格 2 標籤',
                'spec3Label': '規格 3 標籤',
                'spec4Label': '規格 4 標籤',
                'spec1Icon': '規格 1 圖示',
                'spec2Icon': '規格 2 圖示',
                'spec3Icon': '規格 3 圖示',
                'spec4Icon': '規格 4 圖示',
                'trialMode': '試用模式（如：previewFlag）',
                'trialLimit': '試用限制數量（預設 3）',
                'releaseAtMs': '⭐ 新增：發布時間戳（毫秒，Unix timestamp * 1000）',
                'createdAtMs': '⭐ 新增：建立時間戳（毫秒，Unix timestamp * 1000）',
            }
        },
        'FEATURED_LISTS': {
            'columns': [
                'listId',         # ⭐ 必要
                'title',          # ⭐ 必要
                'type',           # ⭐ 必要（productIds / topicIds）
                'ids',            # ⭐ 必要（分號分隔的 ID 列表）
            ],
            'descriptions': {
                'listId': '清單 ID（唯一識別碼）',
                'title': '清單標題',
                'type': '類型（productIds 或 topicIds）',
                'ids': 'ID 列表（多個用分號 ; 分隔）',
            }
        },
        'CONTENT_ITEMS': {
            'columns': [
                'itemId',         # ⭐ 必要
                'productId',      # ⭐ 必要
                'type',           # 選填
                'topicId',        # 選填
                'level',          # 選填
                'levelGoal',      # 選填
                'levelBenefit',   # 選填
                'anchorGroup',    # 選填
                'anchor',         # 選填
                'intent',         # 選填
                'difficulty',     # 選填（數字，預設 1）
                'content',        # 選填
                'wordCount',      # 選填（數字）
                'reusable',       # 選填（true/false，預設 false）
                'sourceType',     # 選填
                'source',         # 選填
                'sourceUrl',      # 選填
                'version',        # 選填
                'pushOrder',      # 選填（數字，推播順序）
                'storageFile',    # 選填
                'seq',            # 選填（數字，預設 0）
                'isPreview',      # 選填（true/false，預設 false）
            ],
            'descriptions': {
                'itemId': '內容項目 ID（唯一識別碼）',
                'productId': '所屬產品 ID',
                'type': '內容類型',
                'topicId': '所屬主題 ID',
                'level': '等級',
                'levelGoal': '等級目標描述',
                'levelBenefit': '等級效益描述',
                'anchorGroup': '錨點群組',
                'anchor': '錨點',
                'intent': '意圖',
                'difficulty': '難度（1-5，預設 1）',
                'content': '內容文字',
                'wordCount': '字數',
                'reusable': '是否可重複使用（true/false，預設 false）',
                'sourceType': '來源類型',
                'source': '來源',
                'sourceUrl': '來源 URL',
                'version': '版本號',
                'pushOrder': '推播順序（數字，Day N）',
                'storageFile': '儲存檔案路徑',
                'seq': '序列號（數字，預設 0）',
                'isPreview': '是否為預覽（true/false，預設 false）',
            }
        }
    }
    
    # 創建 Excel writer
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for sheet_name, sheet_info in sheets_data.items():
            # 創建空 DataFrame
            df = pd.DataFrame(columns=sheet_info['columns'])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f'✅ 創建工作表: {sheet_name}')
            print(f'   欄位數: {len(sheet_info['columns'])}')
            
            # 獲取工作表對象以進行格式化
            worksheet = writer.sheets[sheet_name]
            
            # 設置標題行樣式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            # 標記必要欄位
            required_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            optional_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            # 格式化標題行
            for col_idx, col_name in enumerate(sheet_info['columns'], 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 設置欄寬
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 20
                
                # 判斷是否為必要欄位（根據 check_excel_structure.py 的定義）
                required_fields = {
                    'UI_SEGMENTS': ['segmentId', 'title', 'order', 'mode', 'published'],
                    'TOPICS': ['topicId', 'title', 'published', 'order'],
                    'PRODUCTS': ['productId', 'topicId', 'level'],
                    'FEATURED_LISTS': ['listId', 'title', 'type', 'ids'],
                    'CONTENT_ITEMS': ['itemId', 'productId'],
                }
                
                is_required = col_name in required_fields.get(sheet_name, [])
                if is_required:
                    # 必要欄位：紅色背景
                    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                elif '新增' in sheet_info['descriptions'].get(col_name, ''):
                    # 新增欄位：黃色背景
                    cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
            
            # 添加說明行（第二行）
            desc_row = 2
            for col_idx, col_name in enumerate(sheet_info['columns'], 1):
                cell = worksheet.cell(row=desc_row, column=col_idx)
                description = sheet_info['descriptions'].get(col_name, '')
                cell.value = description
                cell.font = Font(size=9, italic=True, color='666666')
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            # 設置說明行高度
            worksheet.row_dimensions[desc_row].height = 40
            
            print(f'   必要欄位: {len([c for c in sheet_info["columns"] if c in required_fields.get(sheet_name, [])])}')
            print(f'   選填欄位: {len([c for c in sheet_info["columns"] if c not in required_fields.get(sheet_name, [])])}')
            print()
    
    print(f'✅ 新模板已創建: {output_excel}')
    print(f'\n📊 欄位統計:')
    print(f'   UI_SEGMENTS: {len(sheets_data["UI_SEGMENTS"]["columns"])} 個欄位')
    print(f'   TOPICS: {len(sheets_data["TOPICS"]["columns"])} 個欄位')
    print(f'   PRODUCTS: {len(sheets_data["PRODUCTS"]["columns"])} 個欄位（含 2 個新增欄位：releaseAtMs, createdAtMs）')
    print(f'   FEATURED_LISTS: {len(sheets_data["FEATURED_LISTS"]["columns"])} 個欄位')
    print(f'   CONTENT_ITEMS: {len(sheets_data["CONTENT_ITEMS"]["columns"])} 個欄位')
    print(f'\n💡 使用說明:')
    print(f'   1. 打開 {output_excel}')
    print(f'   2. 第一行是欄位名稱（紅色=必要，黃色=新增，藍色=選填）')
    print(f'   3. 第二行是欄位說明')
    print(f'   4. 從第三行開始填入資料')
    print(f'   5. 執行上傳腳本: python3 upload_v3_excel.py --key tools/keys/service-account.json --excel {output_excel}')
    print(f'\n⭐ 新增欄位說明:')
    print(f'   PRODUCTS.releaseAtMs: 發布時間戳（毫秒），用於排序和顯示「本週新泡泡」')
    print(f'   PRODUCTS.createdAtMs: 建立時間戳（毫秒），用於排序和顯示「本週新泡泡」')
    print(f'   時間戳計算方式: int(time.time() * 1000) 或 datetime.now().timestamp() * 1000')

if __name__ == '__main__':
    import sys
    output_file = 'learning_bubble_template_new.xlsx'
    if len(sys.argv) > 1:
        output_file = sys.argv[1]
    
    try:
        create_new_template(output_file)
    except Exception as e:
        print(f'❌ 錯誤: {e}')
        import traceback
        traceback.print_exc()
